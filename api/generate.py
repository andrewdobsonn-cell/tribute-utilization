import sys
import io
import os
import json
import tempfile
from http.server import BaseHTTPRequestHandler
from collections import defaultdict
from datetime import datetime
import re

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError as e:
    pass

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

MKT_TAGS  = ["Boston","NOB","North-of-Boston","MWB","SOB",
              "Chicago","Annapolis","Baltimore","Bethesda","Bel-Air"]
MKT_NAMES = ["Boston","North of Boston","North of Boston","MetroWest Boston",
              "South of Boston","Chicago","Annapolis","Baltimore","Bethesda","Bel-Air"]
MKT_PRI   = ["North-of-Boston","NOB","MWB","SOB","Bethesda",
              "Annapolis","Baltimore","Bel-Air","Chicago","Boston"]
MKT_ORDER = ["Boston","North of Boston","MetroWest Boston","South of Boston",
              "Chicago","Annapolis","Baltimore","Bethesda","Bel-Air","Unknown"]
ST_ORDER  = ["Massachusetts","Illinois","Maryland","Unknown"]
STATE_MAP = {
    "Boston":"Massachusetts","North of Boston":"Massachusetts",
    "MetroWest Boston":"Massachusetts","South of Boston":"Massachusetts",
    "Chicago":"Illinois",
    "Annapolis":"Maryland","Baltimore":"Maryland",
    "Bethesda":"Maryland","Bel-Air":"Maryland",
}
EX_NOTES  = {"office associate","office associate - test",
              "office associate/prn caregiver"}
ABSENCE_T = {"Time Off","PTO - New","Sick Time","Turn Down"}
EXCLUDE_T = {"Expense Reimbursement","Bonus 1 Year (Ann)","Bonus 2&3 Year (Ann)",
             "Bonus 4 Year (Ann)","Bonus 5 - 9 Year (Ann)"}
TYPE_NAMES  = ["Long Hours","Overnight","Short Hours","Live-In"]
DAY_NAMES   = ["Sunday","Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"]
DAY_ABBREVS = ["Su","Mo","Tu","We","Th","Fr","Sa"]

# ─────────────────────────────────────────────────────────────────────────────
# COLORS
# ─────────────────────────────────────────────────────────────────────────────

def rgb(r,g,b): return f"{r:02X}{g:02X}{b:02X}"

NAVY      = rgb(15,38,71)
LBLUE     = rgb(44,74,124)
SLATE     = rgb(68,84,106)
WHITE     = rgb(255,255,255)
OFF_WHITE = rgb(248,250,252)
MED_GRAY  = rgb(210,210,210)
DK_GRAY   = rgb(100,100,100)
GREEN_FILL= rgb(226,239,218)
ORANGE_FL = rgb(252,228,214)
YELLOW_FL = rgb(255,242,204)
BLUE_FL   = rgb(218,227,243)
TYPE_COLORS = [GREEN_FILL, ORANGE_FL, YELLOW_FL, BLUE_FL]

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, name="Arial", italic=False):
    return Font(bold=bold, color=color, size=size, name=name, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

# ─────────────────────────────────────────────────────────────────────────────
# MARKET / STATE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def get_market(tags):
    tags_lower = str(tags).lower()
    for tag in MKT_PRI:
        if tag.lower() in tags_lower:
            idx = next(i for i,t in enumerate(MKT_TAGS)
                      if t.lower() == tag.lower())
            return MKT_NAMES[idx]
    return "Unknown"

def get_state(market):
    return STATE_MAP.get(market, "Unknown")

def tag_has_market(tags):
    tags_lower = str(tags).lower()
    return any(t.lower() in tags_lower for t in MKT_TAGS)

# ─────────────────────────────────────────────────────────────────────────────
# NOTES PARSING
# ─────────────────────────────────────────────────────────────────────────────

DAY_PATTERN = re.compile(
    r'(?:Su|Sun|Sunday|Mo|Mon|Monday|Tu|Tue|Tuesday|We|Wed|Wednesday|'
    r'Th|Thu|Thursday|Fr|Fri|Friday|Sa|Sat|Saturday)',
    re.IGNORECASE
)
DAY_MAP = {
    'su':0,'sun':0,'sunday':0,'mo':1,'mon':1,'monday':1,
    'tu':2,'tue':2,'tuesday':2,'we':3,'wed':3,'wednesday':3,
    'th':4,'thu':4,'thursday':4,'fr':5,'fri':5,'friday':5,
    'sa':6,'sat':6,'saturday':6,
}

def parse_avail_days(notes, section):
    result = {}
    notes_lower = notes.lower()
    if section == "Any":
        for m in DAY_PATTERN.finditer(notes):
            d = DAY_MAP.get(m.group().lower(), -1)
            if d >= 0:
                result[d] = result.get(d, 1.0)
        return result
    kw_map = {
        "Long Hours": ["long hour days","long hours days"],
        "Overnight":  ["overnight days","overnights"],
        "Short Hours":["short hour days","short hours days"],
    }
    keywords = kw_map.get(section, [])
    section_start = -1
    for kw in keywords:
        idx = notes_lower.find(kw)
        if idx >= 0:
            section_start = idx
            break
    if section_start < 0:
        return result
    eq_idx = notes.find('=', section_start)
    if eq_idx < 0:
        return result
    next_section = len(notes)
    for kw in ["long hour","overnight","short hour","live-in","live in"]:
        nxt = notes_lower.find(kw, section_start + 1)
        if nxt > section_start and nxt < next_section:
            next_section = nxt
    segment = notes[eq_idx+1:next_section]
    eod_pat = re.compile(
        r'e/o\s+(' + '|'.join(DAY_MAP.keys()) + r')', re.IGNORECASE)
    for m in eod_pat.finditer(segment):
        d = DAY_MAP.get(m.group(1).lower(), -1)
        if d >= 0:
            result[d] = 0.5
    for m in DAY_PATTERN.finditer(segment):
        d = DAY_MAP.get(m.group().lower(), -1)
        if d >= 0 and d not in result:
            pre = segment[:m.start()].strip().lower()
            if pre.endswith('e/o'):
                result[d] = 0.5
            else:
                result[d] = 1.0
    return result

def pt_from_notes(notes):
    nl = notes.lower()
    if "live-in" in nl or "live in" in nl: return 3
    if "long hour" in nl: return 0
    if "overnight" in nl: return 1
    return 2

def parse_days_str(notes, section):
    days = parse_avail_days(notes, section)
    if not days:
        return ""
    return ", ".join(DAY_ABBREVS[d] for d in sorted(days.keys()))

# ─────────────────────────────────────────────────────────────────────────────
# SCHEDULE PROCESSING
# ─────────────────────────────────────────────────────────────────────────────

def classify_visit(start_time, sched_hours, adj_hours):
    try:
        sh = float(sched_hours) if pd.notna(sched_hours) else 0
        ap = float(adj_hours)   if pd.notna(adj_hours)   else sh
    except:
        sh = 0; ap = 0
    hr = start_time.hour if isinstance(start_time, (datetime, pd.Timestamp)) else 0
    if sh == 24 or (hr == 0 and sh >= 12): return 3
    elif 18 <= hr <= 23 and ap < 16: return 1
    elif 6 <= hr <= 12 and ap >= 8:  return 0
    else: return 2

def get_day_idx(start_time):
    if pd.isna(start_time): return -1
    try:
        if isinstance(start_time, (datetime, pd.Timestamp)):
            return (start_time.weekday() + 1) % 7
    except: pass
    return -1

def process_schedule(df_sc):
    type_hrs    = defaultdict(lambda: [0.0]*4)
    type_vis    = defaultdict(lambda: [0]*4)
    day_type_hrs= defaultdict(lambda: [[0.0]*4 for _ in range(7)])
    pto_day_flags=defaultdict(lambda: [False]*7)
    pto_day_hrs = defaultdict(lambda: [0.0]*7)

    for _, row in df_sc.iterrows():
        cid = str(row.get("Caregiver-Id","")).strip()
        if not cid or cid == "nan": continue
        vt  = str(row.get("Visit Type","")).strip()
        cancelled_raw = str(row.get("Cancelled","")).strip().lower()
        is_cancelled  = cancelled_raw not in ("unset","","nan")
        start_time = row.get("Scheduled Start Time")
        if isinstance(start_time, str):
            try: start_time = pd.to_datetime(start_time)
            except: start_time = None
        day_idx = get_day_idx(start_time)
        try: sh = float(row.get("Scheduled Hours",0) or 0)
        except: sh = 0
        try:
            ap_raw = row.get("Adjusted Payable Hours")
            ap = float(ap_raw) if pd.notna(ap_raw) and str(ap_raw) not in ("","nan") else sh
        except: ap = sh

        if vt in ABSENCE_T:
            if day_idx >= 0:
                pto_day_flags[cid][day_idx] = True
                pto_day_hrs[cid][day_idx]  += min(sh, 10.0)
        elif vt not in EXCLUDE_T and vt != "Travel Time":
            cat = classify_visit(start_time, sh, ap)
            if is_cancelled:
                try:
                    ap = float(ap_raw) if (pd.notna(ap_raw) and
                               str(ap_raw) not in ("","nan") and
                               float(ap_raw) > 0) else 0
                except: ap = 0
            if ap > 0:
                type_hrs[cid][cat]  += ap
                type_vis[cid][cat]  += 1
                if day_idx >= 0:
                    day_type_hrs[cid][day_idx][cat] += ap
    return type_hrs, type_vis, day_type_hrs, pto_day_flags, pto_day_hrs

# ─────────────────────────────────────────────────────────────────────────────
# CAREGIVER PROCESSING
# ─────────────────────────────────────────────────────────────────────────────

def process_caregivers(df_cg, type_hrs, type_vis, day_type_hrs,
                        pto_day_flags, pto_day_hrs):
    active_rows = []
    prn_rows    = []
    pto_rows    = []

    for _, row in df_cg.iterrows():
        notes  = str(row.get("Caregiver Availability Notes","") or "").strip()
        tags   = str(row.get("Tags","") or "").strip()
        cg_id  = str(row.get("Caregiver-Id","") or "").strip()
        if not cg_id or cg_id == "nan": continue
        if notes.lower() in EX_NOTES: continue
        if not tag_has_market(tags) and notes.lower() in ("unset","","nan"): continue
        if "prn" in tags.lower(): continue

        mkt     = get_market(tags)
        state   = get_state(mkt)
        mkt_idx = MKT_ORDER.index(mkt) if mkt in MKT_ORDER else len(MKT_ORDER)-1
        st_idx  = ST_ORDER.index(state) if state in ST_ORDER else len(ST_ORDER)-1

        try:
            desired_raw = row.get("Caregiver Availability Hours Per Week",0)
            desired = float(desired_raw) if pd.notna(desired_raw) and \
                      str(desired_raw) not in ("","Unset","nan") else 0.0
        except: desired = 0.0

        cg_type_hrs = type_hrs.get(cg_id,  [0.0]*4)
        cg_type_vis = type_vis.get(cg_id,  [0]*4)
        cg_day_type = day_type_hrs.get(cg_id, [[0.0]*4 for _ in range(7)])
        cg_pto_flgs = pto_day_flags.get(cg_id, [False]*7)
        cg_pto_hrs  = pto_day_hrs.get(cg_id,  [0.0]*7)

        worked    = sum(cg_type_hrs)
        max_vis   = max(cg_type_vis)
        prim_type = cg_type_vis.index(max_vis) if max_vis > 0 else \
                    (pt_from_notes(notes) if worked == 0 else 2)

        # Live-in conversion
        if prim_type == 3:
            li_worked = sum(min(cg_day_type[d][3] * 1.5, 24.0)
                           for d in range(7) if cg_day_type[d][3] > 0)
            worked = worked - cg_type_hrs[3] + li_worked

        nl        = notes.lower()
        is_li     = "live-in" in nl or "live in" in nl
        has_hourly= "long hour" in nl or "overnight" in nl or "short hour" in nl
        is_mixed  = is_li and has_hourly

        if worked == 0 and is_mixed and prim_type == 3:
            if "long hour" in nl:   prim_type = 0
            elif "overnight" in nl: prim_type = 1
            else:                   prim_type = 2

        pt_name = TYPE_NAMES[prim_type]

        if is_li and prim_type == 3:
            avail_days = {d: 1.0 for d in range(7)}
            daily_rate = 24; pt_name = "Live-In"; desired = 168.0
        elif is_li and is_mixed and prim_type != 3:
            if "long hour" in nl:
                avail_days = parse_avail_days(notes, "Long Hours")
                pt_name = "Live-In / Long Hours"
            else:
                avail_days = parse_avail_days(notes, "Overnight")
                pt_name = "Live-In / Overnight"
            daily_rate = 12
        elif is_li and not is_mixed and prim_type != 3:
            avail_days = {}; daily_rate = 0
        elif prim_type == 0:
            avail_days = parse_avail_days(notes, "Long Hours"); daily_rate = 12
        elif prim_type == 1:
            avail_days = parse_avail_days(notes, "Overnight");  daily_rate = 12
        else:
            avail_days = parse_avail_days(notes, "Any");        daily_rate = 0

        remain_w = sum(w for d,w in avail_days.items() if not cg_pto_flgs[d])

        if is_li and not is_mixed and prim_type != 3:
            adj_cap = worked
        elif daily_rate > 0:
            max_ach = remain_w * daily_rate
            adj_cap = min(max_ach, desired) if max_ach >= desired else max_ach
            adj_cap = max(adj_cap, worked)
        else:
            adj_cap = max(desired, worked)
        adj_cap = max(adj_cap, 0.0)
        util = (worked / adj_cap) if adj_cap > 0 else None

        day_cap = [0.0]*7; day_worked = [0.0]*7; day_avail = [0.0]*7
        if daily_rate > 0:
            for d in range(7):
                wt = avail_days.get(d, 0.0)
                if wt > 0 and not cg_pto_flgs[d]:
                    day_cap[d] = daily_rate * wt; day_avail[d] = wt
                day_worked[d] = cg_day_type[d][prim_type]
            for d in range(7):
                if day_cap[d] == 0 and day_worked[d] > 0:
                    day_cap[d] = daily_rate; day_avail[d] = 1.0

        tc_pto = sum(min(cg_pto_hrs[d], 10.0) for d in range(7))
        tags_lower = tags.lower()
        ts_type = 40 if "tributesecure40" in tags_lower else \
                  30 if "tributesecure30" in tags_lower else 0

        any_days = parse_avail_days(notes, "Any")
        day_display = []
        for d in range(7):
            wt = avail_days.get(d, 0.0)
            if wt == 0.5:          day_display.append("e/o")
            elif d in any_days:    day_display.append("Y")
            elif is_li:            day_display.append("Live-In")
            else:                  day_display.append("N")

        base = {
            "cg_num":    str(row.get("Caregiver Number","") or ""),
            "first":     str(row.get("First Name","") or ""),
            "last":      str(row.get("Last Name","") or ""),
            "market":    mkt, "state": state,
            "status":    str(row.get("Status","") or ""),
            "desig":     str(row.get("Designation","") or ""),
            "svc_type":  pt_name, "desired": desired,
            "adj_cap":   adj_cap, "worked":  worked,
            "visits":    sum(cg_type_vis), "util": util,
            "ld_days":   parse_days_str(notes, "Long Hours"),
            "od_days":   parse_days_str(notes, "Overnight"),
            "sd_days":   parse_days_str(notes, "Short Hours"),
            "is_li":     "Yes" if is_li else "No",
            "day_disp":  day_display, "notes": notes, "tags": tags,
            "hr":        str(row.get("HR Assignee Name","") or ""),
            "mkt_idx":   mkt_idx, "st_idx": st_idx,
            "prim_type": prim_type, "day_cap": day_cap,
            "day_worked":day_worked, "day_avail": day_avail,
            "ts_type":   ts_type, "pto_hrs": tc_pto, "cg_id": cg_id,
        }

        if desired > 0 and adj_cap == 0 and worked == 0:
            pto_rows.append(base)
        elif desired == 0 and worked > 0:
            prn_rows.append({**base, "svc_type": "PRN Worked"})
        elif desired > 0 or worked > 0:
            active_rows.append(base)

    return active_rows, prn_rows, pto_rows

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def set_col_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def style_cell(cell, bg=None, fg="000000", bold=False, size=10,
               h_align="left", v_align="center", wrap=False, bc=None):
    if bg: cell.fill = fill(bg)
    cell.font      = font(bold=bold, color=fg, size=size)
    cell.alignment = align(h=h_align, v=v_align, wrap=wrap)
    if bc: cell.border = border(bc)

def header_row(ws, row_idx, values, bg=SLATE, fg=WHITE, height=18):
    ws.row_dimensions[row_idx].height = height
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=col, value=val)
        style_cell(c, bg=bg, fg=fg, bold=True, size=9,
                   h_align="center" if col > 1 else "left", bc="BBBBBB")

def title_row(ws, text, ncols):
    ws.row_dimensions[1].height = 22
    c = ws.cell(row=1, column=1, value=text)
    c.fill = fill(NAVY); c.font = font(bold=True, color=WHITE, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

# ─────────────────────────────────────────────────────────────────────────────
# TAB WRITERS
# ─────────────────────────────────────────────────────────────────────────────

def write_caregiver_detail(ws, active_rows, prn_rows):
    ws.title = "Caregiver Detail"
    ws.freeze_panes = "A3"
    headers = [
        "CG#","First Name","Last Name","Market","State","Status","Designation",
        "Service Type","Desired Hrs","Adj Capacity","Hrs Worked","Visits","Util %",
        "Long Hour Days","Overnight Days","Short Hour Days","Live-In",
        "Sun","Mon","Tue","Wed","Thu","Fri","Sat","Notes","Tags","HR Assignee"
    ]
    title_row(ws, "Caregiver Utilization Report — Caregiver Detail", len(headers))
    header_row(ws, 2, headers)
    ri = 3
    for rows, gray in [(active_rows, False), (prn_rows, True)]:
        for r in rows:
            bg = MED_GRAY if gray else None
            fg = DK_GRAY  if gray else "000000"
            vals = [r["cg_num"],r["first"],r["last"],r["market"],r["state"],
                    r["status"],r["desig"],r["svc_type"],r["desired"],
                    r["adj_cap"],r["worked"],r["visits"],r["util"],
                    r["ld_days"],r["od_days"],r["sd_days"],r["is_li"]
                    ] + r["day_disp"] + [r["notes"],r["tags"],r["hr"]]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=col, value=val)
                style_cell(c, bg=bg, fg=fg, size=9,
                           h_align="center" if col > 3 else "left", bc="DDDDDD")
            for col in [9,10,11]:
                ws.cell(row=ri, column=col).number_format = "#,##0.0"
            if not gray:
                uc = ws.cell(row=ri, column=13)
                if r["util"] is not None:
                    uc.number_format = "0.0%"
                    if r["util"] == 0:   uc.fill = fill(YELLOW_FL)
                    elif r["util"] > 1:  uc.fill = fill(ORANGE_FL)
                else:
                    uc.number_format = "@"; uc.fill = fill(YELLOW_FL)
            ri += 1
    set_col_widths(ws, {
        "A":8,"B":14,"C":16,"D":18,"E":14,"F":10,"G":10,"H":18,
        "I":10,"J":10,"K":10,"L":7,"M":8,"N":16,"O":16,"P":16,"Q":8,
        "R":5,"S":5,"T":5,"U":5,"V":5,"W":5,"X":5,"Y":40,"Z":30,"AA":18
    })

def write_summary_sheet(ws, tab_name, active_rows, prn_rows, pto_rows,
                         key_field, order_list):
    ws.title = tab_name
    ws.freeze_panes = "A3"
    headers = [("State" if tab_name=="State Summary" else "Market/State"),
               "Service Type","Count","Desired Hrs",
               "Adj Capacity","Hrs Worked","Util %","Over 100%","Zero Util"]
    title_row(ws, f"Caregiver Utilization Report — {tab_name}", len(headers))
    header_row(ws, 2, headers)
    ri = 3
    for key in order_list:
        key_rows = [r for r in active_rows if r[key_field] == key]
        if not key_rows: continue
        tot_cap    = sum(r["adj_cap"] for r in key_rows)
        tot_worked = sum(r["worked"]  for r in key_rows)
        tot_util   = tot_worked/tot_cap if tot_cap > 0 else None
        tot_over   = sum(1 for r in key_rows if r["util"] and r["util"] > 1)
        tot_zero   = sum(1 for r in key_rows if r["util"] is not None and r["util"]==0)
        row_vals   = [key,"Active Caregivers",len(key_rows),
                      sum(r["desired"] for r in key_rows),
                      tot_cap,tot_worked,tot_util,tot_over,tot_zero]
        ws.row_dimensions[ri].height = 16
        for col,val in enumerate(row_vals,1):
            c = ws.cell(row=ri,column=col,value=val)
            style_cell(c,bg=NAVY,fg=WHITE,bold=True,size=9,
                       h_align="center" if col>1 else "left",bc="BBBBBB")
        for col in [4,5,6]:
            ws.cell(row=ri,column=col).number_format="#,##0.0"
        if tot_util is not None:
            ws.cell(row=ri,column=7).number_format="0.0%"
        ri += 1
        for ti,tn in enumerate(TYPE_NAMES):
            tr = [r for r in key_rows if r["prim_type"]==ti]
            if not tr: continue
            tc2   = sum(r["adj_cap"] for r in tr)
            tw    = sum(r["worked"]  for r in tr)
            tu    = tw/tc2 if tc2>0 else None
            tvals = [key,f"  {tn}",len(tr),
                     sum(r["desired"] for r in tr),tc2,tw,tu,
                     sum(1 for r in tr if r["util"] and r["util"]>1),
                     sum(1 for r in tr if r["util"] is not None and r["util"]==0)]
            for col,val in enumerate(tvals,1):
                c = ws.cell(row=ri,column=col,value=val)
                style_cell(c,bg=TYPE_COLORS[ti],fg="1E293B",size=9,
                           h_align="center" if col>1 else "left",bc="DDDDDD")
            for col in [4,5,6]:
                ws.cell(row=ri,column=col).number_format="#,##0.0"
            if tu is not None:
                ws.cell(row=ri,column=7).number_format="0.0%"
                if tu>1:  ws.cell(row=ri,column=7).fill=fill(ORANGE_FL)
                elif tu==0: ws.cell(row=ri,column=7).fill=fill(YELLOW_FL)
            ri += 1
        prn_k = [r for r in prn_rows if r[key_field]==key]
        if prn_k:
            for col in range(1,len(headers)+1):
                c = ws.cell(row=ri,column=col)
                style_cell(c,bg=MED_GRAY,fg=DK_GRAY,size=9,
                           h_align="center" if col>1 else "left",bc="CCCCCC")
            ws.cell(row=ri,column=1).value=key
            ws.cell(row=ri,column=2).value="  PRN Worked"
            ws.cell(row=ri,column=3).value=len(prn_k)
            ws.cell(row=ri,column=6).value=sum(r["worked"] for r in prn_k)
            ws.cell(row=ri,column=6).number_format="#,##0.0"
            ri += 1
        pto_k = [r for r in pto_rows if r[key_field]==key]
        if pto_k:
            tc_counts = {}
            for r in pto_k:
                tn = TYPE_NAMES[r["prim_type"]]
                tc_counts[tn] = tc_counts.get(tn,0)+1
            detail = ", ".join(f"{v} {k}" for k,v in tc_counts.items() if v>0)
            for col in range(1,len(headers)+1):
                c = ws.cell(row=ri,column=col)
                style_cell(c,bg=MED_GRAY,fg=DK_GRAY,size=9,
                           h_align="left",bc="CCCCCC")
            ws.cell(row=ri,column=1).value=key
            ws.cell(row=ri,column=2).value="  On PTO"
            ws.cell(row=ri,column=3).value=len(pto_k)
            ws.cell(row=ri,column=4).value=detail
            ri += 1
        ri += 1
    set_col_widths(ws,{"A":22,"B":20,"C":8,"D":12,"E":12,"F":12,"G":9,"H":10,"I":10})

def write_ts_sheet(ws, active_rows):
    ws.title = "TS Admin Tracker"
    ws.freeze_panes = "A3"
    headers = ["CG#","First Name","Last Name","Market","State",
               "TS Type","Threshold","Hrs Worked","PTO Hrs","Admin Payout"]
    title_row(ws,"TS Admin Tracker",len(headers))
    header_row(ws,2,headers)
    ts_rows = sorted([r for r in active_rows if r["ts_type"]>0],
                     key=lambda r:(r["market"],r["last"],r["first"]))
    ri = 3
    for r in ts_rows:
        threshold = r["ts_type"]
        payout    = max(0.0, threshold - r["worked"] - r["pto_hrs"])
        vals = [r["cg_num"],r["first"],r["last"],r["market"],r["state"],
                f"TS{threshold}",threshold,r["worked"],r["pto_hrs"],payout]
        for col,val in enumerate(vals,1):
            c = ws.cell(row=ri,column=col,value=val)
            style_cell(c,size=9,h_align="center" if col>3 else "left",bc="DDDDDD")
        for col in [7,8,9,10]:
            ws.cell(row=ri,column=col).number_format="#,##0.0"
        if payout > 0:          ws.cell(row=ri,column=10).fill=fill(YELLOW_FL)
        if payout >= threshold*0.5: ws.cell(row=ri,column=10).fill=fill(ORANGE_FL)
        ri += 1
    set_col_widths(ws,{"A":8,"B":14,"C":16,"D":18,"E":14,
                       "F":8,"G":10,"H":10,"I":10,"J":12})

def write_travel_time(ws, df_sc, df_cg):
    ws.title = "Travel Time"
    ws.freeze_panes = "A3"
    headers = ["CG#","First Name","Last Name","Market","State",
               "Day","Date","Gap Start","Gap End","Gap Hrs","Client Pair"]
    title_row(ws,"Travel Time Gap Analysis",len(headers))
    header_row(ws,2,headers)
    cg_lookup = {}
    for _,row in df_cg.iterrows():
        cid = str(row.get("Caregiver-Id","") or "").strip()
        if cid and cid!="nan":
            tags = str(row.get("Tags","") or "")
            cg_lookup[cid] = {
                "cg_num": str(row.get("Caregiver Number","") or ""),
                "first":  str(row.get("First Name","") or ""),
                "last":   str(row.get("Last Name","") or ""),
                "market": get_market(tags),
                "state":  get_state(get_market(tags)),
            }
    exclude_vt = ABSENCE_T | EXCLUDE_T | {
        "Travel Time","Office Meeting","Orientation","Training",
        "Unpaid Training","Shadow Visit","Home Visit","Leadership Development"}
    def is_client_visit(row):
        vt  = str(row.get("Visit Type","")).strip()
        cfn = str(row.get("Client First Name","") or "").lower()
        can = str(row.get("Cancelled","")).strip().lower()
        return can in ("unset","","nan") and vt not in exclude_vt and "office" not in cfn
    qual = df_sc[df_sc.apply(is_client_visit,axis=1)].copy()
    qual["_start"] = pd.to_datetime(qual["Scheduled Start Time"],errors="coerce")
    qual["_end"]   = pd.to_datetime(qual["Scheduled End Time"],  errors="coerce")
    qual["_date"]  = qual["_start"].dt.date
    qual["_cid"]   = qual["Caregiver-Id"].astype(str).str.strip()
    qual = qual.dropna(subset=["_start","_end","_cid"])
    tt = df_sc[df_sc["Visit Type"]=="Travel Time"].copy()
    tt["_start"] = pd.to_datetime(tt["Scheduled Start Time"],errors="coerce")
    tt["_end"]   = pd.to_datetime(tt["Scheduled End Time"],  errors="coerce")
    tt["_cid"]   = tt["Caregiver-Id"].astype(str).str.strip()
    gaps = []
    for (cid,date),grp in qual.groupby(["_cid","_date"]):
        grp = grp.sort_values("_start").reset_index(drop=True)
        for i in range(len(grp)-1):
            a = grp.iloc[i]; b = grp.iloc[i+1]
            gap_mins = (b["_start"]-a["_end"]).total_seconds()/60
            if not (0 < gap_mins <= 120): continue
            cg_tt = tt[tt["_cid"]==cid]
            covered = any(r["_start"]<=a["_end"] and r["_end"]>=b["_start"]
                         for _,r in cg_tt.iterrows())
            if covered: continue
            info = cg_lookup.get(cid,{})
            cl_a = f"{a.get('Client First Name','')} {a.get('Client Last Name','')}".strip()
            cl_b = f"{b.get('Client First Name','')} {b.get('Client Last Name','')}".strip()
            gaps.append({
                "cg_num": info.get("cg_num",""),
                "first":  info.get("first",""),
                "last":   info.get("last",""),
                "market": info.get("market",""),
                "state":  info.get("state",""),
                "day":    a["_start"].strftime("%a"),
                "date":   date,
                "gap_start": a["_end"],
                "gap_end":   b["_start"],
                "gap_hrs":   gap_mins/60,
                "pair":   f"{cl_a} → {cl_b}",
            })
    ri = 3
    for g in gaps:
        vals = [g["cg_num"],g["first"],g["last"],g["market"],g["state"],
                g["day"],g["date"],g["gap_start"],g["gap_end"],g["gap_hrs"],g["pair"]]
        for col,val in enumerate(vals,1):
            c = ws.cell(row=ri,column=col,value=val)
            style_cell(c,size=9,h_align="center" if col>3 else "left",bc="DDDDDD")
        ws.cell(row=ri,column=10).number_format="#,##0.00"
        ri += 1
    if ri==3:
        ws.cell(row=3,column=1).value="No travel time gaps found this week."
    set_col_widths(ws,{"A":8,"B":14,"C":16,"D":18,"E":14,
                       "F":6,"G":12,"H":18,"I":18,"J":8,"K":35})

def write_roster_sheet(ws, active_rows):
    ws.title = "Caregiver Roster"
    ws.freeze_panes = "A3"
    headers = ["Caregiver","State","Market","Day","Service Type",
               "Adj Capacity","Hrs Worked (Week)","Week Util %"]
    title_row(ws,"Caregiver Roster — Day View",len(headers))
    header_row(ws,2,headers)
    ri = 3
    for mkt in MKT_ORDER:
        for r in [r for r in active_rows
                  if r["market"]==mkt and r["prim_type"] in (0,1,3)]:
            for d in range(7):
                if r["day_cap"][d]==0 and r["day_worked"][d]==0: continue
                cap=r["day_cap"][d]; wrk=r["day_worked"][d]
                util=wrk/cap if cap>0 else None
                vals=[f"{r['first']} {r['last']}",r["state"],r["market"],
                      DAY_NAMES[d],r["svc_type"],cap,wrk,util]
                for col,val in enumerate(vals,1):
                    c=ws.cell(row=ri,column=col,value=val)
                    style_cell(c,size=9,h_align="center" if col>1 else "left",bc="DDDDDD")
                ws.cell(row=ri,column=6).number_format="#,##0.0"
                ws.cell(row=ri,column=7).number_format="#,##0.0"
                if util is not None:
                    ws.cell(row=ri,column=8).number_format="0.0%"
                    if util>1:  ws.cell(row=ri,column=8).fill=fill(ORANGE_FL)
                    elif util==0: ws.cell(row=ri,column=8).fill=fill(YELLOW_FL)
                ri += 1
    set_col_widths(ws,{"A":22,"B":14,"C":18,"D":12,"E":18,"F":12,"G":14,"H":10})

def write_day_breakdown(ws, active_rows):
    ws.title = "Day Breakdown"
    ws.freeze_panes = "A3"
    headers = ["Market","Service Type","Sun","Mon","Tue","Wed","Thu","Fri","Sat"]
    title_row(ws,"Day Breakdown by Market",len(headers))
    header_row(ws,2,headers)
    ri = 3
    for mkt in MKT_ORDER:
        mkt_rows = [r for r in active_rows if r["market"]==mkt]
        if not mkt_rows: continue
        for ti in [0,1,3]:
            tr = [r for r in mkt_rows if r["prim_type"]==ti]
            if not tr: continue
            d_caps  = [sum(r["day_cap"][d]    for r in tr) for d in range(7)]
            d_works = [sum(r["day_worked"][d] for r in tr) for d in range(7)]
            if all(c==0 and w==0 for c,w in zip(d_caps,d_works)): continue
            vals = [mkt,"Overall"]+[
                f"{d_works[d]:.0f}/{d_caps[d]:.0f}" if d_caps[d]>0 else ""
                for d in range(7)]
            for col,val in enumerate(vals,1):
                c=ws.cell(row=ri,column=col,value=val)
                style_cell(c,bg=LBLUE,fg=WHITE,bold=True,size=9,
                           h_align="center" if col>1 else "left",bc="BBBBBB")
            ri += 1
            for r in sorted(tr,key=lambda r:(r["last"],r["first"])):
                vals=[f"{r['first']} {r['last']}",r["svc_type"]]+[
                    f"{r['day_worked'][d]:.1f}/{r['day_cap'][d]:.1f}"
                    if r["day_cap"][d]>0 else "" for d in range(7)]
                for col,val in enumerate(vals,1):
                    c=ws.cell(row=ri,column=col,value=val)
                    style_cell(c,bg=TYPE_COLORS[ti],size=9,
                               h_align="center" if col>1 else "left",bc="DDDDDD")
                ri += 1
        ri += 1
    set_col_widths(ws,{"A":22,"B":18,"C":10,"D":10,"E":10,
                       "F":10,"G":10,"H":10,"I":10})

# ─────────────────────────────────────────────────────────────────────────────
# VERCEL HANDLER
# ─────────────────────────────────────────────────────────────────────────────

class handler(BaseHTTPRequestHandler):

    def do_GET(self):
        self.send_response(200)
        self.send_header("Content-Type","text/plain")
        self.end_headers()
        self.wfile.write(b"Tribute Utilization API is running.")

    def do_POST(self):
        try:
            content_type = self.headers.get("Content-Type","")
            content_length = int(self.headers.get("Content-Length",0))
            body = self.rfile.read(content_length)

            # Parse multipart form data
            import cgi
            environ = {
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE":   content_type,
                "CONTENT_LENGTH": str(content_length),
            }
            fs = cgi.FieldStorage(
                fp=io.BytesIO(body),
                environ=environ,
                keep_blank_values=True
            )

            cg_item = fs["caregiver"]
            sc_item = fs["schedule"]

            cg_bytes = cg_item.file.read()
            sc_bytes = sc_item.file.read()

            df_cg = pd.read_csv(io.StringIO(cg_bytes.decode("utf-8","replace")),
                                dtype=str, low_memory=False)
            df_sc = pd.read_csv(io.StringIO(sc_bytes.decode("utf-8","replace")),
                                dtype=str, low_memory=False)

            for col in ["Caregiver Availability Hours Per Week"]:
                if col in df_cg.columns:
                    df_cg[col] = pd.to_numeric(df_cg[col], errors="coerce")
            for col in ["Scheduled Hours","Adjusted Payable Hours"]:
                if col in df_sc.columns:
                    df_sc[col] = pd.to_numeric(df_sc[col], errors="coerce")

            type_hrs,type_vis,day_type_hrs,pto_flags,pto_hrs = process_schedule(df_sc)
            active_rows,prn_rows,pto_rows = process_caregivers(
                df_cg,type_hrs,type_vis,day_type_hrs,pto_flags,pto_hrs)

            wb = Workbook()
            wb.remove(wb.active)
            write_caregiver_detail(wb.create_sheet("Caregiver Detail"), active_rows, prn_rows)
            write_summary_sheet(wb.create_sheet("Market Summary"),
                "Market Summary",active_rows,prn_rows,pto_rows,"market",MKT_ORDER)
            write_summary_sheet(wb.create_sheet("State Summary"),
                "State Summary",active_rows,prn_rows,pto_rows,"state",ST_ORDER)
            write_day_breakdown(wb.create_sheet("Day Breakdown"), active_rows)
            write_roster_sheet(wb.create_sheet("Caregiver Roster"), active_rows)
            write_ts_sheet(wb.create_sheet("TS Admin Tracker"), active_rows)
            write_travel_time(wb.create_sheet("Travel Time"), df_sc, df_cg)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            xlsx_bytes = output.read()

            self.send_response(200)
            self.send_header("Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition",
                "attachment; filename=CaregiverUtilization.xlsx")
            self.send_header("Content-Length", str(len(xlsx_bytes)))
            self.end_headers()
            self.wfile.write(xlsx_bytes)

        except Exception as e:
            import traceback
            err = traceback.format_exc()
            self.send_response(500)
            self.send_header("Content-Type","text/plain")
            self.end_headers()
            self.wfile.write(err.encode())
